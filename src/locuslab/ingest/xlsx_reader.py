"""XLSX content reader producing one span per non-empty cell."""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from locuslab.ingest.ids import make_span_id
from locuslab.ingest.reader_base import ReaderResult
from locuslab.models import ParserWarning, ParserWarningCode, Span, SpanLocation, SpanLocationKind

PARSER_NAME = f"openpyxl:{openpyxl.__version__}"


def read_xlsx(path: Path, *, document_id: str) -> ReaderResult:
    """Extract one span per non-empty cell, preserving sheet/A1 address."""
    if not path.exists():
        raise FileNotFoundError(f"XLSX file not found: {path}")

    relative = path.name
    try:
        workbook = openpyxl.load_workbook(str(path), data_only=True, read_only=False)
    except Exception as exc:
        return ReaderResult(
            spans=(),
            warnings=(
                ParserWarning(
                    code=ParserWarningCode.EXTRACTION_UNREADABLE_FILE,
                    message=f"openpyxl could not open file: {exc.__class__.__name__}",
                    path=relative,
                ),
            ),
            parser=PARSER_NAME,
        )

    spans: list[Span] = []
    warnings: list[ParserWarning] = []
    running_index = 0

    for sheet_name in workbook.sheetnames:
        try:
            worksheet = workbook[sheet_name]
        except Exception as exc:
            warnings.append(
                ParserWarning(
                    code=ParserWarningCode.EXTRACTION_PARTIAL_CONTENT,
                    message=f"openpyxl could not read sheet: {exc.__class__.__name__}",
                    path=relative,
                    location=f"sheet={sheet_name}",
                )
            )
            continue
        sheet_spans, running_index = _spans_from_sheet(
            worksheet=worksheet,
            document_id=document_id,
            running_index=running_index,
        )
        spans.extend(sheet_spans)

    workbook.close()

    if not spans:
        warnings.append(
            ParserWarning(
                code=ParserWarningCode.EXTRACTION_EMPTY_DOCUMENT,
                message="XLSX opened cleanly but produced no content spans.",
                path=relative,
            )
        )

    return ReaderResult(spans=tuple(spans), warnings=tuple(warnings), parser=PARSER_NAME)


def _spans_from_sheet(
    *,
    worksheet: Worksheet,
    document_id: str,
    running_index: int,
) -> tuple[list[Span], int]:
    spans: list[Span] = []
    sheet_name = worksheet.title

    headers: dict[int, str] = {}
    header_row_iter = iter(worksheet.iter_rows(min_row=1, max_row=1, values_only=False))
    header_row = next(header_row_iter, None)
    if header_row is not None:
        for cell in header_row:
            value = _cell_text(cell.value)
            if value and cell.column is not None:
                headers[cell.column] = value

    for row in worksheet.iter_rows(values_only=False):
        row_values = [_cell_text(cell.value) for cell in row]
        if not any(row_values):
            continue
        for cell, text in zip(row, row_values, strict=True):
            if not text:
                continue
            if cell.column is None or cell.row is None:
                continue
            column_index: int = cell.column
            row_index: int = cell.row
            address = f"{get_column_letter(column_index)}{row_index}"
            label = f"{sheet_name}:{address}"
            if row_index == 1:
                section: str | None = f"{sheet_name}:header"
            else:
                header_value = headers.get(column_index)
                if header_value is not None:
                    column_letter = get_column_letter(column_index)
                    section = f"{sheet_name}:header:{column_letter}={header_value}"
                else:
                    section = None
            location = SpanLocation(
                kind=SpanLocationKind.TABLE_CELL, index=running_index, label=label
            )
            spans.append(
                Span(
                    span_id=make_span_id(document_id=document_id, location=location, text=text),
                    document_id=document_id,
                    location=location,
                    text=text,
                    section=section,
                )
            )
            running_index += 1

    return spans, running_index


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()
