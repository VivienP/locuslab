"""Build findings.xlsx via openpyxl.

One required sheet `Findings` mirroring findings.csv columns plus three
empty adjudication columns (reviewer / review_notes / resolution) ready for
human RA/QA fill-in. Optional `Summary` sheet with severity and finding_type
counts.

Byte-determinism is not asserted by the Phase 5 tests (xlsx is a zip with
internal mtimes); content order is stable because we iterate sorted
sequences.
"""

from __future__ import annotations

import datetime as dt
from collections.abc import Sequence
from pathlib import Path
from typing import cast

from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from locuslab.models import Finding, FindingSeverity

XLSX_FINDINGS_COLUMNS: tuple[str, ...] = (
    "eco_id",
    "severity",
    "checker_id",
    "finding_type",
    "affected_object_ids",
    "evidence",
    "remediation_hint",
    "adjudication_state",
    "reviewer",
    "review_notes",
    "resolution",
)

XLSX_FINDINGS_COLUMN_WIDTHS: dict[str, int] = {
    "eco_id": 22,
    "severity": 14,
    "checker_id": 36,
    "finding_type": 40,
    "affected_object_ids": 50,
    "evidence": 70,
    "remediation_hint": 70,
    "adjudication_state": 18,
    "reviewer": 16,
    "review_notes": 40,
    "resolution": 24,
}

# Fixed timestamp to stabilise visible workbook metadata across reruns.
# Does not make the .xlsx byte-equal (zip internal mtimes still drift) but
# keeps the visible `properties.created` / `modified` deterministic.
_FIXED_TIMESTAMP = dt.datetime(2026, 1, 1, 0, 0, 0)


def _stabilise_workbook_properties(wb: Workbook) -> None:
    wb.properties.creator = "LocusLab"
    wb.properties.lastModifiedBy = "LocusLab"
    wb.properties.created = _FIXED_TIMESTAMP
    wb.properties.modified = _FIXED_TIMESTAMP
    wb.properties.title = "LocusLab Findings"


def _write_findings_sheet(wb: Workbook, findings: Sequence[Finding]) -> None:
    ws = cast(Worksheet, wb.active)
    ws.title = "Findings"
    ws.append(list(XLSX_FINDINGS_COLUMNS))
    header_font = Font(bold=True)
    for cell in ws[1]:
        cast(Cell, cell).font = header_font
    for col_idx, name in enumerate(XLSX_FINDINGS_COLUMNS, start=1):
        header_cell = cast(Cell, ws.cell(row=1, column=col_idx))
        ws.column_dimensions[header_cell.column_letter].width = (
            XLSX_FINDINGS_COLUMN_WIDTHS[name]
        )
    for f in sorted(findings, key=lambda x: x.eco_id):
        ws.append(
            [
                f.eco_id,
                f.severity.value,
                f.checker_id,
                f.finding_type,
                ";".join(f.affected_object_ids),
                f.evidence,
                f.remediation_hint,
                f.adjudication_state.value,
                "",
                "",
                "",
            ]
        )


def _write_summary_sheet(wb: Workbook, findings: Sequence[Finding]) -> None:
    ws = cast(Worksheet, wb.create_sheet(title="Summary"))
    ws.append(["Section", "Key", "Count"])
    for cell in ws[1]:
        cast(Cell, cell).font = Font(bold=True)

    severity_counts: dict[str, int] = {sev.value: 0 for sev in FindingSeverity}
    type_counts: dict[str, int] = {}
    for f in findings:
        severity_counts[f.severity.value] += 1
        type_counts[f.finding_type] = type_counts.get(f.finding_type, 0) + 1

    for sev in (
        FindingSeverity.CRITICAL.value,
        FindingSeverity.MAJOR.value,
        FindingSeverity.MINOR.value,
        FindingSeverity.INFORMATIONAL.value,
    ):
        ws.append(["by_severity", sev, severity_counts[sev]])
    for ftype in sorted(type_counts):
        ws.append(["by_finding_type", ftype, type_counts[ftype]])


def write_findings_xlsx(findings: Sequence[Finding], path: Path) -> None:
    """Write findings.xlsx with a Findings sheet and a Summary sheet."""
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    _stabilise_workbook_properties(wb)
    _write_findings_sheet(wb, findings)
    _write_summary_sheet(wb, findings)
    wb.save(str(path))
